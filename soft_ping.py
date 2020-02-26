from scapy.layers.inet import IP, ICMP
from scapy.sendrecv import sr1
from time import sleep
import sys
import random
import ipaddress
import openpyxl

"""
Python3 script for running low-risk ip address scan using icmp 'ping'.
Sends only one packet at a time. For network or IP ranges, randomizes
target IP. Tries max 4 pings. Randomizes interval of sending packets.
Returns CSV of active IPs and inactive IPs for further testing.
"""
def main(argv):
    user_target = get_target(argv[0])
    print(str(user_target))
    string_list = format_target(user_target)
    #Build a recon csv workbook
    #""" Future feature
    recon_wb = openpyxl.Workbook()
    sheet = recon_wb.get_sheet_by_name('Sheet')
    sheet['A1'].value = 'Target'
    sheet['B1'].value = str(argv[0])
    sheet['A2'].value = 'All addresses'
    sheet['C2'].value = 'Active addresses'
    sheet['B2'].value = 'Unresponsive addresses'
    #"""
    #loop through generated list of target IPs
    #"""
    for index, packet in enumerate(string_list, start = 0):

        #input packet dest in 'All addreses column'
        sheet.cell(row = (index + 3), column = 1, value = str(packet.dst))
        delay = random.random()
        sleep(delay)
        ping_packet = (IP(dst=packet.dst)/ICMP())
        ans = sr1(ping_packet, timeout = 4, verbose=0)
        if not (ans is None):
            print(ans.src + " is online.")
            print(str(ans.src))
            sheet.cell(row = (index + 3), column = 3, value = str(ans.src))
            string_list.remove(packet)
        else:
            print("Timeout waiting for %s " %packet.dst)
            sheet.cell(row = (index + 3), column = 2, value = str(packet.dst))
    recon_wb.save(filename = 'test_workbook.xls')

#"""
#    recon_wb.save(filename = 'test_workbook.xls')


def get_target(ip):
#Determines network or host, checks format, loops until proper address provided
        need_address = True
        #ip = unicode(ip)
        while(need_address == True):
            print("\nIs target a network? (Y/N)")
            user_response = str(input())

            if user_response == 'Y' or user_response == 'y':
                #try:
                addr = ipaddress.IPv4Network(ip)
                need_address = False
                #except:
                    #print("Wrong format/incorrect address. Try CIDR format")
            elif user_response == 'N' or user_response == 'n':
                #try:
                addr = ipaddress.IPv4Address(ip)
                need_address = False
                #except:
                    #print("\nWrong format/incorrect address. Try CIDR format")
            else:
                print("\nProvide correct response. (Y/N)")
        return addr



def format_target(target):
#formats target into useable scapy strings
    addr_string_list = []
    scapy_string_list = []
    print(str(type(target)))
    if str(type(target)) == "<class 'ipaddress.IPv4Network'>":
        print("\nentered network branch\n")
        net_mask = target.netmask
        b_cast = target.broadcast_address
        for host in target.hosts():
            if host != b_cast:
                addr_string_list.append(str(host))
                #print(str(host))
        #lets get weird
        random.shuffle(addr_string_list)
        """
        print("\nAfter randomization\n")
        for host in addr_string_list:
            print(str(host))
        """
        for host in addr_string_list:
            scapy_target_address = IP(dst=str(host))
            scapy_string_list.append(scapy_target_address)
        #print("After scapy-zation\n")
        """
        for host in scapy_string_list:
            print(host.summary())
        """
    elif str(type(target)) == "<class 'ipaddress.IPv4Address'>":
        scapy_string_list.append(IP(dst=str(target)))
    else:
        print("Address translation failed. Try again")

    """
    for item in scapy_string_list:
        print(item.summary())
    """
    return scapy_string_list




if __name__ == '__main__':
    main(sys.argv[1:])
